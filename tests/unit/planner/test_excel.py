"""Tests for planner/excel.py (TDD)."""

from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import load_workbook

from factorydaemon.planner.engine import plan
from factorydaemon.planner.excel import write_excel_report
from factorydaemon.storage.norms import NormStorage


def _norms() -> NormStorage:
    return NormStorage(
        {
            "cut": 120.0,
            "sew": 180.0,
            "pack": 60.0,
        }
    )


def test_write_excel_report_creates_file(tmp_path: Path):
    norms = _norms()
    result = plan(
        demands={"cut": 10, "sew": 5, "pack": 20},
        priorities={"cut": 3, "sew": 2, "pack": 1},
        norms=norms,
        max_positions_per_worker=10,
    )
    output = tmp_path / "plan.xlsx"
    report = write_excel_report(result, output, warnings=["demo warning"])
    assert report.path.exists()


def test_excel_sheets_exist(tmp_path: Path):
    norms = _norms()
    result = plan({"cut": 1}, {"cut": 1}, norms)
    output = tmp_path / "plan.xlsx"
    write_excel_report(result, output)

    wb = load_workbook(output)
    assert set(wb.sheetnames) == {"План", "Сводка", "Предупреждения"}


def test_plan_sheet_headers_and_rows(tmp_path: Path):
    norms = _norms()
    result = plan(
        demands={"cut": 10, "sew": 5},
        priorities={"cut": 1, "sew": 1},
        norms=norms,
        max_positions_per_worker=10,
    )
    output = tmp_path / "plan.xlsx"
    write_excel_report(result, output)

    wb = load_workbook(output)
    ws = wb["План"]
    headers = [ws.cell(row=1, column=c).value for c in range(1, 9)]
    assert headers == [
        "Работник",
        "Загрузка (ч)",
        "Доля смены (%)",
        "Позиция 1",
        "Кол-во 1",
        "Позиция 2",
        "Кол-во 2",
        "Позиция 3",
    ]
    # Worker 1 row.
    assert ws.cell(row=2, column=1).value == 1
    # Time and share columns are in hours / percent.
    assert ws.cell(row=2, column=2).value <= 10.0
    assert ws.cell(row=2, column=3).value <= 100.0
    # Positions appear across columns.
    row_positions = {ws.cell(row=2, column=c).value for c in range(4, 8, 2)}
    assert row_positions == {"cut", "sew"}


def test_plan_sheet_single_position_row(tmp_path: Path):
    norms = _norms()
    result = plan({"cut": 1}, {"cut": 1}, norms)
    output = tmp_path / "plan.xlsx"
    write_excel_report(result, output)

    wb = load_workbook(output)
    ws = wb["План"]
    assert ws.cell(row=2, column=4).value == "cut"
    assert ws.cell(row=2, column=5).value == 1.0
    # Unused position columns are left blank.
    assert ws.cell(row=2, column=6).value is None


def test_summary_sheet_metrics(tmp_path: Path):
    norms = _norms()
    result = plan(
        demands={"cut": 10, "sew": 5, "pack": 20},
        priorities={"cut": 3, "sew": 2, "pack": 1},
        norms=norms,
        max_positions_per_worker=10,
    )
    output = tmp_path / "plan.xlsx"
    write_excel_report(result, output)

    wb = load_workbook(output)
    ws = wb["Сводка"]
    metrics = {ws.cell(row=r, column=1).value: ws.cell(row=r, column=2).value for r in range(2, 8)}
    assert metrics["Количество работников"] == 1
    assert metrics["Длительность смены (ч)"] == 10.0
    assert metrics["Лимит позиций на работника"] == 10
    assert metrics["Рекомендуемое количество работников"] == 1

    # Per-worker summary header.
    w_row = 9
    headers = [ws.cell(row=w_row, column=c).value for c in range(1, 6)]
    assert headers == ["Работник", "Позиций", "Загрузка (ч)", "Загрузка (%)", "Осталось (ч)"]

def test_warnings_sheet_with_warnings(tmp_path: Path):
    norms = _norms()
    result = plan({"cut": 1}, {"cut": 1}, norms)
    output = tmp_path / "plan.xlsx"
    write_excel_report(result, output, warnings=["low utilization", "missing norm X"])

    wb = load_workbook(output)
    ws = wb["Предупреждения"]
    assert ws.cell(row=1, column=1).value == "Предупреждение"
    assert ws.cell(row=2, column=1).value == "low utilization"
    assert ws.cell(row=3, column=1).value == "missing norm X"


def test_warnings_sheet_empty(tmp_path: Path):
    norms = _norms()
    result = plan({"cut": 1}, {"cut": 1}, norms)
    output = tmp_path / "plan.xlsx"
    write_excel_report(result, output)

    wb = load_workbook(output)
    ws = wb["Предупреждения"]
    assert ws.cell(row=2, column=1).value == "Нет предупреждений"


if __name__ == "__main__":
    pytest.main([__file__, "-v"])
