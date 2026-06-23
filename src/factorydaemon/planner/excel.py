"""Excel report generator for a shift plan."""

from __future__ import annotations

import os
from collections.abc import Iterable
from dataclasses import dataclass
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from factorydaemon.planner.engine import PlanResult


@dataclass(frozen=True)
class ExcelReport:
    path: Path


def _set_header(ws: Worksheet, row: int, columns: list[str]) -> None:
    for col_idx, value in enumerate(columns, start=1):
        cell = ws.cell(row=row, column=col_idx, value=value)
        cell.font = Font(bold=True)


def _autosize_columns(ws: Worksheet) -> None:
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                length = len(str(cell.value))
                if length > max_length:
                    max_length = length
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max(max_length + 2, 8), 50)


def write_excel_report(
    plan_result: PlanResult,
    output_path: str | os.PathLike[str],
    warnings: Iterable[str] | None = None,
) -> ExcelReport:
    """Write an Excel file with three sheets: План, Сводка, Предупреждения.

    The `План` sheet is laid out horizontally: one row per worker, with columns
    ``Позиция 1``, ``Кол-во 1``, ``Позиция 2``, ``Кол-во 2`` and so on, up to
    the configured ``max_positions_per_worker`` limit.
    """
    wb = Workbook()

    # Remove default sheet, create named ones.
    wb.remove(wb.active)
    plan_ws = wb.create_sheet("План")
    summary_ws = wb.create_sheet("Сводка")
    warnings_ws = wb.create_sheet("Предупреждения")

    # Sheet: План — one row per worker, positions across columns.
    max_positions = max(
        (w.positions_count for w in plan_result.workers),
        default=0,
    )
    max_positions = max(max_positions, plan_result.max_positions_per_worker)

    header = ["Работник", "Загрузка (ч)", "Доля смены (%)"]
    for idx in range(1, max_positions + 1):
        header.extend([f"Позиция {idx}", f"Кол-во {idx}"])
    _set_header(plan_ws, 1, header)

    row = 2
    for worker in plan_result.workers:
        plan_ws.cell(row=row, column=1, value=worker.index + 1)
        used_hours = worker.used_seconds / 3600
        plan_ws.cell(row=row, column=2, value=round(used_hours, 2))
        share = (
            worker.used_seconds / plan_result.shift_seconds * 100
            if plan_result.shift_seconds > 0
            else 0.0
        )
        plan_ws.cell(row=row, column=3, value=round(share, 2))
        col = 4
        for load in worker.loads:
            plan_ws.cell(row=row, column=col, value=load.position)
            plan_ws.cell(row=row, column=col + 1, value=round(load.units, 2))
            col += 2
        row += 1
    _autosize_columns(plan_ws)

    # Sheet: Сводка
    _set_header(summary_ws, 1, ["Метрика", "Значение"])
    total_seconds = plan_result.total_seconds
    worker_count = plan_result.worker_count
    shift_seconds = plan_result.shift_seconds
    summary_rows = [
        ("Количество работников", worker_count),
        ("Рекомендуемое количество работников", plan_result.required_workers or worker_count),
        ("Длительность смены (ч)", round(shift_seconds / 3600, 2)),
        ("Лимит позиций на работника", plan_result.max_positions_per_worker),
        ("Общая трудоёмкость (ч)", round(total_seconds / 3600, 2)),
        ("Средняя загрузка (%)", round(plan_result.utilization * 100, 2)),
    ]
    for idx, (metric, value) in enumerate(summary_rows, start=2):
        summary_ws.cell(row=idx, column=1, value=metric)
        summary_ws.cell(row=idx, column=2, value=value)

    # Per-worker summary rows.
    w_row = len(summary_rows) + 3
    summary_ws.cell(row=w_row, column=1, value="Работник")
    summary_ws.cell(row=w_row, column=2, value="Позиций")
    summary_ws.cell(row=w_row, column=3, value="Загрузка (ч)")
    summary_ws.cell(row=w_row, column=4, value="Загрузка (%)")
    summary_ws.cell(row=w_row, column=5, value="Осталось (ч)")
    for cell in summary_ws[w_row]:
        cell.font = Font(bold=True)
    w_row += 1
    for worker in plan_result.workers:
        used = worker.used_seconds
        cap = worker.capacity_seconds
        summary_ws.cell(row=w_row, column=1, value=worker.index + 1)
        summary_ws.cell(row=w_row, column=2, value=worker.positions_count)
        summary_ws.cell(row=w_row, column=3, value=round(used / 3600, 2))
        summary_ws.cell(row=w_row, column=4, value=round(used / cap * 100, 2) if cap > 0 else 0.0)
        summary_ws.cell(row=w_row, column=5, value=round((cap - used) / 3600, 2))
        w_row += 1
    _autosize_columns(summary_ws)

    # Sheet: Предупреждения
    _set_header(warnings_ws, 1, ["Предупреждение"])
    warning_list = list(warnings or [])
    if not warning_list:
        warnings_ws.cell(row=2, column=1, value="Нет предупреждений")
    else:
        for idx, text in enumerate(warning_list, start=2):
            warnings_ws.cell(row=idx, column=1, value=text)
    _autosize_columns(warnings_ws)

    path = Path(output_path)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return ExcelReport(path=path)
